#==================================================================================================================================================================
#   Schema Configuration Tool
#   Brandon Hearley - LMS
#   6/4/2024
#
#   PURPOSE: Create a web app (using streamlit) to filter data collected and stored on the NASA
#            GRC Lab Infrastructure and generate a Word Report
#
#==================================================================================================================================================================
# SETUP
# Import the necessary modules to run the app and set paths

# Import Modules
import time
import json
import streamlit as st
from openpyxl import load_workbook

# Set Home Directory
raw_template = "/mount/src/schemaconfiguration/Raw_Template.json"
analysis_template = "/mount/src/schemaconfiguration/Analysis_Template.json"

#==================================================================================================================================================================
# GENERAL INFORMATION
# Set the web app general information not edited by the user

# Set the page configuration
st.set_page_config(layout="wide")

# Create the Title
st.title("PyMILab Schema Configuration Manager")

#==================================================================================================================================================================
# SCHEMA SELECTION
# Generate a new configuation file or load one from a previous save

if "excel_flag" not in st.session_state:
    # Initialize Session State
    
    

    # Create Instructions
    instruct1 = st.empty()
    instruct1.markdown("Upload either a new Excel (.xlsx) file to configure a new schema or a previous Configuration (.json) file to " +  
                       "load a previous configuration. For new configurations, enter a unique schema configuration name.")

    

    # Create File Uploader Button
    file = st.empty()
    filename = file.file_uploader('Upload a Excel Schema or Configuration File', type = ['xlsx','json'],
                            accept_multiple_files = False, key = "file")

    if filename != None:
        if 'xlsx' in st.session_state['file'].name:
            st.session_state['excel_flag'] = 1
            st.session_state['json_flag'] = 0
        else:
            st.session_state['excel_flag'] = 0
            st.session_state['json_flag'] = 1

        if st.button('Configure Schema'):
            temp=1 # Just rerun code

else:
    if st.session_state['excel_flag'] == 1:
        st.session_state['excel_flag'] = 2
        
        # Read The Excel File and Get MI Attributes
        wb = load_workbook(st.session_state['file'], data_only=True, read_only=True)
        # Get List of Sheets
        Sheets = []
        i = 0
        while wb.sheetnames[i] != 'Data':
            Sheets.append(wb.sheetnames[i])
            i=i+1
        Sheets.append('Data')

        # Get List of Attributes
        Atts = {'Single Value':{},
                'Functional':{},
                'Tabular':{}}

        # Read the tabular and functional data attributes
        for i in range(len(Sheets)-1):
            # Open the sheet
            ws = wb[Sheets[i]]

            # Get the attribute name
            att_name = ws.cell(row=4,column=2).value

            # Determine if the attribute is functional or tabular
            # -- 0 = functional
            # -- 1 = tabular
            att_flag = 0
            if ws.cell(row=7,column = 3).value == 'Row Number':
                att_flag = 1
                row_num = 7

            # Get Functional Data Information
            if att_flag == 0:
                # Get X and Y Names
                x_name = ws.cell(row = 8,column=3).value
                y_name = ws.cell(row = 8,column=4).value

                # Check for Units
                x_att = x_name
                x_unit = None
                if x_name[-1] == ')':
                    idx = x_name.index("(")
                    x_att = x_name[:idx-1]
                    x_unit = x_name[idx+1:len(x_name)-1]

                y_att = y_name
                y_unit = None
                if y_name[-1] == ')':
                    idx = y_name.index("(")
                    y_att = y_name[:idx-1]
                    y_unit = y_name[idx+1:len(y_name)-1]

                Atts['Functional'][att_name] = {'Variables':[x_att, y_att],
                                                'Units':[x_unit, y_unit]}

            # Get Tabular Data Information
            else:
                # Get the editable column names (don't include row number)
                cols = []
                units = []
                col_num = 4
                while ws.cell(row=row_num,column=col_num).value != None:
                    # Get the color of cell
                    clr = ws.cell(row=row_num,column=col_num).fill.start_color.index 
                    if clr == 'FFFFFF00':
                        # Check for associated units
                        row_name = ws.cell(row=row_num,column=col_num).value
                        row_att = row_name
                        row_unit = None
                        if row_name[-1] == ')':
                            idx = row_name.index("(")
                            row_att = row_name[:idx-1]
                            row_unit = row_name[idx+1:len(row_name)-1]
                            temp=1

                        cols.append(row_att)
                        units.append(row_unit)
                    col_num = col_num+1
                Atts['Tabular'][att_name] = {'Columns':cols,
                                                'Units':units}

        # Get The Single Value Attributes
        ws = wb['Data']

        for k in range(10, ws.max_row+1):
            # Check the color for a header
            if ws.cell(row=k,column=3).fill.start_color.index == 'FFFFFFFF' and ws.cell(row=k,column=3).value != None:
                Atts['Single Value'][ws.cell(row=k,column=3).value]= ws.cell(row=k,column=4).value

        # Set the Excel Flag
        st.session_state['Atts'] = Atts

    elif st.session_state['json_flag'] == 1:
        st.session_state['json_flag'] = 2

        # Load the Previous Configuration
        Prev_Config = json.load(st.session_state['file'])
        st.session_state['Config'] = Prev_Config

        # Get the Atts List
        st.session_state['Atts'] = Prev_Config['Atts']


    # Load the Raw and Analysis Template File
    f = open(raw_template)
    Raw = json.load(f)
    st.session_state['Raw'] = Raw

    f = open(analysis_template)
    Analysis = json.load(f)
    st.session_state['Analysis'] = Analysis

    # Load Atts
    Atts = st.session_state['Atts']
    Raw = st.session_state['Raw']
    Analysis = st.session_state['Analysis']

    # Initialize the Configuration JSON
    if 'Config' not in st.session_state:
        Config = {}

        # -- Single Attributes
        Config['Single Value'] = {}
        single_atts = list(Atts['Single Value'].keys())
        for i in range(len(single_atts)):
            Config['Single Value'][single_atts[i]] = ''

        # -- Functional Attributes
        Config['Functional'] = {}
        func_atts = list(Atts['Functional'].keys())
        for i in range(len(func_atts)):
            Config['Functional'][func_atts[i]] = {}
            Config['Functional'][func_atts[i]]['X'] = ''
            Config['Functional'][func_atts[i]]['Y'] = ''

        # -- Tabular Attributes
        Config['Tabular'] = {}
        tab_atts = list(Atts['Tabular'].keys())
        for i in range(len(tab_atts)):
            Config['Tabular'][tab_atts[i]] = {}
            Config['Tabular'][tab_atts[i]]['GrantaCols'] = Atts['Tabular'][tab_atts[i]]['Columns']
            temp = []
            for j in range(len(Atts['Tabular'][tab_atts[i]]['Columns'])):
                temp.append('')
            Config['Tabular'][tab_atts[i]]['PyCols'] = temp

        # -- Save Atts for future use
        Config['Atts'] = st.session_state['Atts']

        st.session_state['Config'] = Config

    with st.expander('Single Value Attributes'):
        # Get List of Schema Attributes
        atts = list(Atts['Single Value'].keys())
        Config = st.session_state['Config']

         # Get List of all JSON Attributes
        JSON_atts = ['']
        # -- Raw Data
        Raw_cat = list(Raw.keys())
        for j in range(len(Raw_cat)):
            Raw_att = list(Raw[Raw_cat[j]].keys())
            for k in range(len(Raw_att)):
                if Raw[Raw_cat[j]][Raw_att[k]]['Type'] == 'point'or Raw[Raw_cat[j]][Raw_att[k]]['Type'] == 'string':
                    att_name = Raw_cat[j] + ' - ' + Raw_att[k]
                    JSON_atts.append(att_name)

        # -- Analysis Data
        Analysis_cat = list(Analysis.keys())
        for j in range(len(Analysis_cat)):
            Analysis_att = list(Analysis[Analysis_cat[j]].keys())
            for k in range(len(Analysis_att)):
                if Analysis[Analysis_cat[j]][Analysis_att[k]]['Type'] == 'point'or Analysis[Analysis_cat[j]][Analysis_att[k]]['Type'] == 'string':
                    att_name = Analysis_cat[j] + ' - ' + Analysis_att[k]
                    JSON_atts.append(att_name)

        # Create the table
        single_grid = st.empty()
        grid = single_grid.columns(2)

        for i in range(len(atts)):
            with grid[0]:
                if i == 0:
                    st.text_input('Database Attribute',value = atts[i], key = f'single_val_a_{i}')
                else:
                    st.text_input('Database Attribute',value = atts[i], key = f'single_val_a_{i}',label_visibility = "collapsed")
            with grid[1]:
                if i == 0:
                    st.selectbox('Py MI Lab Attribute', JSON_atts, index = JSON_atts.index(Config['Single Value'][atts[i]]), key=f'single_val_b_{i}')
                else:
                    st.selectbox('Py MI Lab Attribute', JSON_atts, index = JSON_atts.index(Config['Single Value'][atts[i]]), key=f'single_val_b_{i}',label_visibility = "collapsed")

        # Save the data
        Config = st.session_state['Config']
        for i in range(len(atts)):
            Config['Single Value'][atts[i]] = st.session_state[f'single_val_b_{i}']
        st.session_state['Config'] = Config

        # Save Single JSON atts
        st.session_state['single_json'] = JSON_atts

    with st.expander('Functional Attributes'):
        # Get List of Schema Attributes
        atts = list(Atts['Functional'].keys())

        # Get List of all JSON Attributes
        JSON_atts = ['']
        # -- Raw Data
        Raw_cat = list(Raw.keys())
        for j in range(len(Raw_cat)):
            Raw_att = list(Raw[Raw_cat[j]].keys())
            for k in range(len(Raw_att)):
                if Raw[Raw_cat[j]][Raw_att[k]]['Type'] == 'point array':
                    att_name = Raw_cat[j] + ' - ' + Raw_att[k]
                    JSON_atts.append(att_name)

         # -- Analysis Data
        Analysis_cat = list(Analysis.keys())
        for j in range(len(Analysis_cat)):
            Analysis_att = list(Analysis[Analysis_cat[j]].keys())
            for k in range(len(Analysis_att)):
                if Analysis[Analysis_cat[j]][Analysis_att[k]]['Type'] == 'point array':
                    att_name = Analysis_cat[j] + ' - ' + Analysis_att[k]
                    JSON_atts.append(att_name)


        # Create the table
        single_grid = st.empty()
        grid = single_grid.columns(3)

        for i in range(len(atts)):
            with grid[0]:
                if i == 0:
                    st.text_input('Database Attribute',value = atts[i], key = f'func_a_{i}')
                else:
                    st.text_input('Database Attribute',value = atts[i], key = f'func_a_{i}',label_visibility = "collapsed")
            with grid[1]:
                if i == 0:
                    st.selectbox('X - Py MI Lab Attribute', JSON_atts, index = JSON_atts.index(Config['Functional'][atts[i]]['X']), key=f'func_b_{i}')
                else:
                    st.selectbox('X - Py MI Lab Attribute', JSON_atts, index = JSON_atts.index(Config['Functional'][atts[i]]['X']), key=f'func_b_{i}',label_visibility = "collapsed")
            with grid[2]:
                if i == 0:
                    st.selectbox('Y - Py MI Lab Attribute', JSON_atts, index = JSON_atts.index(Config['Functional'][atts[i]]['Y']), key=f'func_c_{i}')
                else:
                    st.selectbox('Y - Py MI Lab Attribute', JSON_atts, index = JSON_atts.index(Config['Functional'][atts[i]]['Y']), key=f'func_c_{i}',label_visibility = "collapsed")

        # Save the data
        Config = st.session_state['Config']
        for i in range(len(atts)):
            Config['Functional'][atts[i]]['X'] = st.session_state[f'func_b_{i}']
            Config['Functional'][atts[i]]['Y'] = st.session_state[f'func_c_{i}']
        st.session_state['Config'] = Config


    if 'tab_exp' not in st.session_state:
            st.session_state['tab_exp'] = False
    else:
        st.session_state['tab_exp'] = True

    JSON_atts = ['']
    # -- Raw Data
    Raw_cat = list(Raw.keys())
    for j in range(len(Raw_cat)):
        Raw_att = list(Raw[Raw_cat[j]].keys())
        for k in range(len(Raw_att)):
            if Raw[Raw_cat[j]][Raw_att[k]]['Type'] != 'dict':
                att_name = Raw_cat[j] + ' - ' + Raw_att[k]
                JSON_atts.append(att_name)

    # -- Analysis Data
    Analysis_cat = list(Analysis.keys())
    for j in range(len(Analysis_cat)):
        Analysis_att = list(Analysis[Analysis_cat[j]].keys())
        for k in range(len(Analysis_att)):
            if Analysis[Analysis_cat[j]][Analysis_att[k]]['Type'] != 'dict':
                att_name = Analysis_cat[j] + ' - ' + Analysis_att[k]
                JSON_atts.append(att_name)

    def update_tab():
        with st.expander('Tabular Attributes', expanded = st.session_state['tab_exp']):
            # Get List of Schema Attributes
            atts = list(Atts['Tabular'].keys())

            if "ct" not in st.session_state:
                st.session_state["ct"] = 0

            # Get the max number of columns
            if "max_col" not in st.session_state:
                max_col = 0
                for i in range(len(atts)):
                    if len(Atts['Tabular'][atts[i]]['Columns']) > max_col:
                        max_col = len(Atts['Tabular'][atts[i]]['Columns']) 
                st.session_state["max_col"] = max_col
                
            # Create a Select Box for the different tabular attributes
            tab_att_opt = st.selectbox('Select the tabular attribute',atts, key='tab_att_opt')

            if 'prev_opt' not in st.session_state:
                st.session_state['prev_opt'] = tab_att_opt

            if "col_names" not in st.session_state:
                st.session_state["col_names"] = {}

            # Initialize the table
            if "tab_init" not in st.session_state:
                st.session_state["tab_init"] = True

            # Get the attribute
            att_name = st.session_state["tab_att_opt"]

            # Get the Current Data
            Config = st.session_state['Config']
            GrantaCols = Config['Tabular'][att_name]['GrantaCols']
            PyCols = Config['Tabular'][att_name]['PyCols']
            
            col_vals = []
            new_vals = []

            if st.session_state['prev_opt'] != st.session_state['tab_att_opt']:
                st.session_state["ct"] = st.session_state["ct"]+1

            tab_cols = st.columns(2)
            D = st.session_state["col_names"]

            for i in range(len(GrantaCols)):
                col_vals.append('')
                new_vals.append('')

                with tab_cols[0]:
                    D["var1_" + str(i)] = st.empty()
                    if i == 0:
                        col_vals[i] = D["var1_" + str(i)].text_input('Database Attribute',value = GrantaCols[i], key = f'tab_a_{st.session_state["ct"]+i}')
                    else:
                        col_vals[i] = D["var1_" + str(i)].text_input('Database Attribute',value = GrantaCols[i], key = f'tab_a_{st.session_state["ct"]+i}', label_visibility="collapsed")
                with tab_cols[1]:
                    D["var2_" + str(i)]= st.empty()

                    # Get index
                    if PyCols[i] == None:
                        idx = None
                    else:
                        idx = JSON_atts.index(PyCols[i])
                    
                    if i == 0:
                        new_vals[i] = D["var2_" + str(i)].selectbox('Py MI Lab Attribute',JSON_atts,index = idx, key = f'tab_b_{st.session_state["ct"]+i}')
                    else:
                        new_vals[i] = D["var2_" + str(i)].selectbox('Database Attribute',JSON_atts, index = idx, key = f'tab_b_{st.session_state["ct"]+i}', label_visibility="collapsed")

                st.session_state["col_names"] = D            
                st.session_state['change_opt'] = False

            if st.session_state['prev_opt'] != st.session_state['tab_att_opt']:
                st.session_state['prev_opt'] = st.session_state['tab_att_opt']


            # Store the Data
            Config['Tabular'][att_name]['PyCols'] = []
            for j in range(len(new_vals)):
                Config['Tabular'][att_name]['PyCols'].append(new_vals[j])

            st.session_state['Config'] = Config

    # Update Tabular Attributes
    update_tab()

    with st.expander('Record Placement'):
        #Re-open Config
        Config = st.session_state['Config']

        if st.session_state['json_flag'] == 2:
            st.session_state['json_flag'] = 3
            if "Placement" in list(Config.keys()):
                num_val = len(list(Config["Placement"].keys()))
        else:
            num_val = 0     

        
        # Create the number of levels
        num_lev = st.number_input('Number of Folder Levels', value = num_val, min_value = 0, step = 1, key = 'num_lev')
    
        def create_cond_table(m,n):
            grid_sec = st.columns([0.035, 0.215, 0.05, 0.2, 0.25, 0.25])
            if st.session_state[f'folder_lev_b_{m}'] != 1:
                with grid_sec[0]:
                    st.text_input('If',value= 'IF', disabled = True, key = f'folder_sec_a_{m}_{n}', label_visibility="collapsed")
                with grid_sec[1]:
                    st.selectbox('Conditional Attribute', st.session_state['single_json'], index=None, placeholder = "Select the conditional attribute", key = f'folder_sec_b_{m}_{n}', label_visibility="collapsed")
                with grid_sec[2]:
                    st.text_input('Eq',value= 'EQ :', disabled = True, key = f'folder_sec_c_{m}_{n}', label_visibility="collapsed")
                with grid_sec[3]:
                    st.text_input('Eq',value= None, placeholder = 'Enter the conditional value', key = f'folder_sec_d_{m}_{n}', label_visibility="collapsed")
            with grid_sec[4]:
                st.selectbox('Conditional Attribute', st.session_state['single_json'], index= None, placeholder = "Select the naming attribute and format", key = f'folder_sec_e_{m}_{n}', label_visibility="collapsed")
            with grid_sec[5]:
                st.text_input('Eq',value= '[attribute]', key = f'folder_sec_f_{m}_{n}', label_visibility="collapsed")
            
        def create_folder_table(m):
            grid_lev = st.columns([0.8,0.2])
            with grid_lev[0]:
                if m == 0:
                    st.text_input("Folder Level", value = "Level " + str(m+1), disabled = True, key = f'folder_lev_a_{m}')
                else:
                    st.text_input("Folder Level", value = "Level " + str(m+1), disabled = True, key = f'folder_lev_a_{m}', label_visibility = "collapsed")
            with grid_lev[1]:
                if m == 0:
                    st.number_input('Conditions', value = 1, min_value = 0, step = 1, key = f'folder_lev_b_{m}')
                else:
                    st.number_input('Conditions', value = 1, min_value = 0, step = 1, key = f'folder_lev_b_{m}', label_visibility = "collapsed")

            
            for n in range(st.session_state[f'folder_lev_b_{m}']): 
                create_cond_table(m,n)


        if st.session_state['num_lev'] != None:
            for m in range(st.session_state['num_lev']):
                create_folder_table(m)
            st.markdown('''---''')

        # Store the data in the Config
        if st.session_state['num_lev'] != None:
            Config = st.session_state['Config']
            Config['Placement'] = {}
            for m in range(st.session_state['num_lev']):
                Config['Placement']['Level ' + str(m+1)] = []
                for n in range(st.session_state[f'folder_lev_b_{m}']):
                    new_vec = []
                    if f'folder_sec_a_{m}_{n}' in st.session_state:
                        new_vec.append(st.session_state[f'folder_sec_a_{m}_{n}'])
                    else:
                        new_vec.append(None)
                    if f'folder_sec_b_{m}_{n}' in st.session_state:
                        new_vec.append(st.session_state[f'folder_sec_b_{m}_{n}'])
                    else:
                        new_vec.append(None)
                    if f'folder_sec_c_{m}_{n}' in st.session_state:
                        new_vec.append(st.session_state[f'folder_sec_c_{m}_{n}'])
                    else:
                        new_vec.append(None)
                    if f'folder_sec_d_{m}_{n}' in st.session_state:
                        new_vec.append(st.session_state[f'folder_sec_d_{m}_{n}'])
                    else:
                        new_vec.append(None)
                    if f'folder_sec_e_{m}_{n}' in st.session_state:
                        new_vec.append(st.session_state[f'folder_sec_e_{m}_{n}'])
                    else:
                        new_vec.append(None)
                    if f'folder_sec_f_{m}_{n}' in st.session_state:
                        new_vec.append(st.session_state[f'folder_sec_f_{m}_{n}'])
                    else:
                        new_vec.append(None)
                    Config['Placement']['Level ' + str(n+1)].append(new_vec)

    # Create the config file
    json_string = json.dumps(st.session_state['Config'])

    #st.json(json_string, expanded=True)
    
    st.download_button(
        label="Download Configuration File",
        file_name="New_Schema_Config.json",
        mime="application/json",
        data=json_string,
    )

            





  


